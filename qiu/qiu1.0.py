from tkinter import filedialog, Tk, Button, Label, Entry, StringVar
import tkinter.messagebox
from openpyxl import load_workbook, Workbook
import threading
from time import sleep


class GUI():

    def __init__(self, root):

        self.data_phones = []
        self.tager_phones = []
        self.results = []
        self.data_path = ""
        self.tager_path = ""
        self.data_value = StringVar()
        self.tager_value = StringVar()
        self.flag = 0
        self.text_first = StringVar()
        self.text_mid = StringVar()
        self.text_last = StringVar()
        self.initGUI(root)

    def initGUI(self, root):
        self.root = root
        self.root.title("秋-v1.0")
        self.root.geometry("340x140+500+300")
        self.root.resizable = False
        Label(root, text="秋-去重工具").grid(row=0, column=3)

        Label(root, text="去重文件：").grid(row=1, column=1)
        Entry(root, textvariable=self.data_value,
              width=30).grid(row=1, column=3)
        Button(root, text="选择文件", command=self.data_select).grid(row=1, column=4)

        Label(root, text="数据文件：").grid(row=2, column=1)
        Entry(root, textvariable=self.tager_value,
              width=30).grid(row=2, column=3)
        Button(root, text="选择文件", command=self.tager_select).grid(
            row=2, column=4)
        Button(root, text="开始", width=10, command=lambda: self.begin(
            self.data_path, self.tager_path)).grid(row=3, column=3)
        Label(root, textvariable=self.text_first).grid(
            row=4, column=1, columnspan=2)
        Label(root, textvariable=self.text_last).grid(
            row=4, column=3, columnspan=3)
        root.mainloop()

    def data_select(self):
        self.data_path = self.selectPath()
        self.data_value.set(self.data_path)

    def tager_select(self):
        self.tager_path = self.selectPath()
        self.tager_value.set(self.tager_path)

    def selectPath(self):
        path = filedialog.askopenfilenames(
            title="请选择文件", filetypes=[("excel file", "*.xlsx *.xls *.txt")])
        return path

    def begin(self, datapath, tagerpath):
        self.read_excel(datapath, 0)
        self.read_excel(tagerpath, 1)
        threading.Thread(target=self.compare).start()

    def read_excel(self, path, num):
        threads = []
        if path == "":
            tkinter.messagebox.showwarning("警告", "您未选择文件，请选择要去重的文件")
        else:
            for filename in path:
                threads.append(threading.Thread(
                    target=self.get_phone, args=(filename, num)))
            for thread in threads:
                thread.setDaemon(True)
                thread.start()

    def get_phone(self, filename, num):
        self.flag += 1
        phone_list = []
        if filename.endswith(".txt"):
            with open(filename, "r") as f:
                for line in f:
                    phone_num = line.replace("\n", "")
                    if num == 0:
                        if phone_num not in self.data_phones:
                            self.data_phones.append(phone_num)
                    else:
                        if phone_num not in self.tager_phones:
                            self.tager_phones.append(phone_num)
        else:
            book = load_workbook(filename=filename, read_only=True)
            table = book.active
            nrows = table.rows
            for row in nrows:
                phone_num = row[0].value
                if num == 0:
                    if phone_num not in self.data_phones:
                        self.data_phones.append(phone_num)
                else:
                    if phone_num not in self.tager_phones:
                        self.tager_phones.append(phone_num)
        if num == 0:
            for i in phone_list:
                if i not in self.data_phones:
                    self.data_phones.append(i)
        else:
            for i in phone_list:
                if i not in self.tager_phones:
                    self.tager_phones.append(i)
        self.flag -= 1

    def compare(self):
        while True:
            if self.flag == 0:
                self.text_first.set("文件去重中")
                count = 0
                for value in self.tager_phones:
                    count += 1
                    self.text_last.set("已处理条数：" + str(count))
                    if value not in self.data_phones:
                        self.results.append(value)
                if len(self.results) != 0:
                    self.write_excel()
                break
            else:
                sleep(0.2)
                self.text_first.set("读取文件中")
                self.text_last.set(
                    "已读取条数：" + str(len(self.data_phones) + len(self.tager_phones)))

    def write_excel(self):
        save_file = filedialog.asksaveasfilename(
            title=u'保存文件', filetypes=[("excel file", "*.xlsx *.xls")])
        if save_file == "":
            tkinter.messagebox.showwarning("警告", "您未选择保存的文件，将不会保存结果")
        else:
            wb = Workbook()
            ws = wb.active
            index = 1
            for phone in self.results:
                ws.cell(row=index, column=1).value = phone
                index += 1
            wb.save(save_file)
            wb.close()
            tkinter.messagebox.showinfo("完成", "文件已去重完毕，结果存放在 "+save_file)


if __name__ == "__main__":
    root = Tk()
    myGUI = GUI(root)
